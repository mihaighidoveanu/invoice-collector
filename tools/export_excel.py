"""Standalone Excel export utility.

Reads a report_{month}.json produced by report_builder and writes the .xlsx
the accountant expects.

Usage:
    python tools/export_excel.py invoices/report_2025-03.json
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Row fill colours
_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_invoices_sheet(ws, results: list[dict]) -> None:
    headers = [
        "Vendor",
        "Date",
        "Amount",
        "Currency",
        "Status",
        "Email Subject",
        "Attachment Path",
        "Failure Reason",
    ]
    ws.append(headers)

    for result in results:
        tx = result["transaction"]
        status = result.get("status", "")
        row = [
            tx.get("vendor", ""),
            tx.get("date", ""),
            tx.get("amount", ""),
            tx.get("currency", ""),
            status,
            result.get("email_subject", ""),
            str(result.get("attachment_path") or ""),
            result.get("failure_reason") or "",
        ]
        ws.append(row)

        fill = _GREEN if status == "found" else _RED
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = fill

    _autofit(ws)


def _build_normalizations_sheet(ws, normalizations: list[dict]) -> None:
    ws.append(["Raw Bank Text", "Normalized Name", "LLM Confidence Note"])
    for norm in normalizations:
        ws.append(
            [
                norm.get("raw_description", ""),
                norm.get("normalized_name", ""),
                norm.get("confidence_note", ""),
            ]
        )
    _autofit(ws)


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def export(report_path: Path, output_path: Path | None = None) -> Path:
    """Convert a JSON pipeline report to an Excel workbook.

    Args:
        report_path: Path to report_{month}.json.
        output_path: Override output .xlsx path. Defaults to same directory.

    Returns:
        Path to the written .xlsx file.
    """
    data = json.loads(report_path.read_text(encoding="utf-8"))

    wb = Workbook()
    ws_invoices = wb.active
    ws_invoices.title = "Invoices"
    _build_invoices_sheet(ws_invoices, data.get("results", []))

    ws_norm = wb.create_sheet("Vendor Normalizations")
    _build_normalizations_sheet(ws_norm, data.get("vendor_normalizations", []))

    if output_path is None:
        output_path = report_path.with_suffix(".xlsx")

    wb.save(output_path)
    print(f"Exported: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python tools/export_excel.py <report.json> [output.xlsx]")
        sys.exit(1)

    report = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    export(report, out)
